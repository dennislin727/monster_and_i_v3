# -*- coding: utf-8 -*-
"""ARCHITECTURE 三槽寵物／跟隨節奏／卡牆 idle／開局種子 → Monster_DevLog_v4.xlsx (02/03/07/100) + docs 快照."""
from __future__ import annotations

import datetime
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Monster_DevLog_v4.xlsx"
ARCH = ROOT / "ARCHITECTURE.md"
SYNC_DATE = datetime.date(2026, 3, 30)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def arch_toc(md: str) -> str:
    lines = md.splitlines()
    toc_lines = []
    for i, ln in enumerate(lines, 1):
        if ln.startswith("## "):
            toc_lines.append(f"L{i}: {ln[3:].strip()}")
    return "\n".join(toc_lines)


def find_sheet(wb, prefix: str) -> str:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    raise KeyError(prefix)


def last_used_row(ws, max_col: int = 12) -> int:
    last = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None and v != "":
                last = r
                break
    return last


def main() -> None:
    import openpyxl

    md = ARCH.read_text(encoding="utf-8")
    toc = arch_toc(md)
    snapshot = (
        f"《怪物與我》ARCHITECTURE.md 同步快照 {SYNC_DATE.isoformat()}（三槽寵物跟隨／ARCHITECTURE 修訂）\n\n"
        f"【章節索引】\n{toc}\n\n---\n"
        "【本批重點】三槽編隊與 GlobalBalance PET_PARTY_SLOT*/TRAIL_LAG、無存檔種子 STARTER_PET_PATHS "
        "可重複同模板（暱稱·序號）、PetCompanion 卡牆 idle（_motion_intent_vel／_seek_point／dot）、"
        "Phase 7 頭飾敘述改多槽、Phase 11 補 user://monster_and_i_save_v1.json、寵物出戰策略定調。\n\n"
        "（全文見 repo ARCHITECTURE.md；GlobalBalance 槽位數值以 scratch 為準。）"
    )
    if len(snapshot) > 32000:
        snapshot = snapshot[:31900] + "\n…(截斷)"

    docs_snap = ROOT / "docs" / "_ARCHITECTURE_sync_snapshot_for_devlog.txt"
    docs_snap.parent.mkdir(parents=True, exist_ok=True)
    docs_snap.write_text(snapshot, encoding="utf-8")
    print("Snapshot:", docs_snap)

    if not XLSX.is_file():
        print("WARN: 找不到", XLSX.name, "— 已僅更新 docs 快照。")
        return

    wb = openpyxl.load_workbook(XLSX)
    s02 = find_sheet(wb, "02_")
    s03 = find_sheet(wb, "03_")
    s07 = find_sheet(wb, "07_")
    s100 = find_sheet(wb, "100_")

    summary = (
        "ARCHITECTURE 同步：三槽寵物跟隨節奏（GlobalBalance SLOT1/2 FOLLOW_MULT、TRAIL_LAG_SEC）與 DOCUMENT；"
        "PetCompanion 卡牆仍 run 修復（意圖速度、朝目標 dot、STUCK_*）；"
        "PetManager 開局種子可重複同 .tres、暱稱·序號、有存檔不種子；"
        "Phase 7 頭飾「多槽已落地」、Phase 11 存檔路徑 user://、寵物出戰策略去「進行中」改定調。"
    )

    r02 = last_used_row(wb[s02]) + 1
    ws02 = wb[s02]
    ws02.cell(r02, 1, SYNC_DATE.isoformat())
    ws02.cell(r02, 2, "ARCHITECTURE 三槽寵物／跟隨／卡牆／種子")
    ws02.cell(r02, 3, summary)
    ws02.cell(
        r02,
        4,
        "GlobalBalance.gd；PetCompanion.gd；PetManager.gd；ARCHITECTURE.md Phase 4/7/11",
    )
    ws02.cell(r02, 5, "是")
    bullets = [
        "【節奏】槽2/3 慢跑＋麵包屑延遲；同 PetResource.follow_speed_mult 仍可疊槽位倍率。",
        "【卡牆】勿以 lerp 後 velocity 判斷意圖；_motion_intent_vel + _seek_point 前進分量。",
        "【種子】STARTER_PET_PATHS 重複路徑→三隻個體；無 JSON 才種子。",
        "【文件】docs/_ARCHITECTURE_sync_snapshot_for_devlog.txt 已更新。",
    ]
    for i, b in enumerate(bullets, start=1):
        ws02.cell(r02 + i, 3, b)

    r03 = last_used_row(wb[s03]) + 1
    ws03 = wb[s03]
    ws03.cell(r03, 1, "P1")
    ws03.cell(r03, 2, "寵物編隊跟隨體感定調（槽位倍率／麵包屑／卡牆動畫，2026-03-30）")
    ws03.cell(
        r03,
        3,
        "ARCHITECTURE 與 GlobalBalance 已落值；公開測試後可微調常數或補槽位 FOLLOW_ARRIVE 實驗",
    )
    ws03.cell(
        r03,
        4,
        "ARCHITECTURE.md；PetCompanion.gd；GlobalBalance.gd；PetManager STARTER_PET_PATHS",
    )
    ws03.cell(r03, 5, "延續：多寵路徑尋飛／障礙繞路若仍穿模再議")
    ws03.cell(r03, 6, "04：本批含 PET_PARTY_* 與 PET_FOLLOW 相關平衡")

    r07 = last_used_row(wb[s07]) + 1
    ws07 = wb[s07]
    ws07.cell(r07, 1, SYNC_DATE)
    ws07.cell(r07, 2, "寵物三槽＋ARCHITECTURE 修訂")
    ws07.cell(
        r07,
        3,
        "多槽出戰體感與卡牆 idle 修復已入版並入 ARCHITECTURE；"
        "開局三隻同源種子利測試；DevLog 02/03/07/100 本批附加。",
    )
    ws07.cell(
        r07,
        5,
        "GlobalBalance PET_PARTY_SLOT*；docs 快照；與 PlayerController 全管線對齊列為靈感牆後續。",
    )

    r100 = last_used_row(wb[s100]) + 1
    ws100 = wb[s100]
    ws100.cell(r100, 1, SYNC_DATE.isoformat())
    ws100.cell(r100, 2, "寵物跟隨：可選與主角同套 move 管線或槽位 FOLLOW_ARRIVE 偏移")
    ws100.cell(r100, 3, f"ARCHITECTURE 同步 {SYNC_DATE.isoformat()}")
    ws100.cell(
        r100,
        4,
        "現行以動畫／意圖閾值＋槽位 lag 滿足體感；若仍穿牆再評估 Navigation 或簡易避障",
    )
    ws100.cell(r100, 5, "與 obstacles／y_sort 同層碰撞現已使用")
    ws100.cell(r100, 6, "—")
    ws100.cell(r100, 7, "三槽長距跟跑迴歸")
    ws100.cell(r100, 8, "觀察中")

    wb.save(XLSX)
    print("Saved:", XLSX)
    print("02 block start:", r02)
    print("03 row:", r03)
    print("07 row:", r07)
    print("100 row:", r100)


if __name__ == "__main__":
    main()
