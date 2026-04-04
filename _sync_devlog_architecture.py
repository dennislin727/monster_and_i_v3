# -*- coding: utf-8 -*-
"""Sync ARCHITECTURE.md into Monster_DevLog_v4.xlsx sheets 02, 03, 07, 100 (Phase 9 MVP 落地備援)."""
from __future__ import annotations

import datetime
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Monster_DevLog_v4.xlsx"
ARCH = ROOT / "ARCHITECTURE.md"

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SYNC_DATE = datetime.date(2026, 3, 29)


def arch_toc_and_next(md: str) -> tuple[str, str]:
    lines = md.splitlines()
    toc_lines = []
    for i, ln in enumerate(lines, 1):
        if ln.startswith("## "):
            toc_lines.append(f"L{i}: {ln[3:].strip()}")
    toc = "\n".join(toc_lines)
    next_block = []
    capture = False
    for ln in lines:
        if ln.strip() == "## 下一階段（文件錨點，供新對話接手）":
            capture = True
            next_block.append(ln)
            continue
        if capture:
            if ln.startswith("### Phase 8 UI") and len(next_block) > 5:
                break
            next_block.append(ln)
    return toc, "\n".join(next_block)


def find_sheet(wb: openpyxl.Workbook, prefix: str) -> str:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    raise KeyError(prefix)


def last_used_row(ws: openpyxl.worksheet.worksheet.Worksheet, max_col: int = 12) -> int:
    last = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None and v != "":
                last = r
                break
    return last


def main() -> None:
    md = ARCH.read_text(encoding="utf-8")
    toc, next_stage = arch_toc_and_next(md)
    snapshot = (
        f"《怪物與我》ARCHITECTURE.md 同步快照 {SYNC_DATE.isoformat()}\n\n"
        f"【章節索引】\n{toc}\n\n---\n{next_stage}\n\n"
        f"（全文見 repo 根目錄 ARCHITECTURE.md；含「湖畔關卡地圖與場景編排（實作定版）」換關儀式感協議。"
        f"04 數值中心本次無變更。）"
    )
    if len(snapshot) > 32000:
        snapshot = snapshot[:31900] + "\n…(截斷)"

    docs_snap = ROOT / "docs" / "_ARCHITECTURE_sync_snapshot_for_devlog.txt"
    docs_snap.parent.mkdir(parents=True, exist_ok=True)
    docs_snap.write_text(snapshot, encoding="utf-8")

    wb = openpyxl.load_workbook(XLSX)
    s02 = find_sheet(wb, "02_")
    s03 = find_sheet(wb, "03_")
    s07 = find_sheet(wb, "07_")
    s100 = find_sheet(wb, "100_")

    summary_main = (
        "ARCHITECTURE 增「湖畔關卡（LakeSideLevel）地圖與場景編排（實作定版，2026-03-29）」："
        "Background＋TerrainCollision 北中南 CollisionPolygon2D；LakeSideLevelRoot 有底圖則隱 TerrainMap；"
        "MarkersPropSpawner call_deferred 撒点至 level_container、prop_data 寫 MonsterBase.data；"
        "ForegroundCanopyHoist 樹冠改掛 LevelContainer；MonsterBase perform_ghost_dash 分段 move_and_collide；"
        "換關保留 UILayer 儀式感（區域名／採收鈕／黑幕節奏協議）；下一張家園獨立場景＋作物嘟嘟；"
        "§5 已知待修：寵物穿牆、對話後採收鈕、進出家園搖桿、道具NPC 無影、動畫粒子 polish。"
        "Phase 10 表列／層級規則／下一階段第0項已對齊；常見地雷增 _ready add_child busy。"
    )

    r02 = last_used_row(wb[s02]) + 1
    ws02 = wb[s02]
    ws02.cell(r02, 1, SYNC_DATE.isoformat())
    ws02.cell(r02, 2, "ARCHITECTURE.md")
    ws02.cell(r02, 3, summary_main)
    ws02.cell(
        r02,
        4,
        "LakeSideLevel 地圖定版；MarkersPropSpawner；ForegroundCanopyHoist；換關儀式感；下一階段家園",
    )
    ws02.cell(r02, 5, "是")
    bullets = [
        "【地圖】LakeSideLevel：Background、TerrainCollision North/Center/South、HomesteadZone/Crops、ScatteredRocks+Slimes。",
        "【腳本】LakeSideLevelRoot 有 texture 則 TerrainMap.hide；MarkersPropSpawner deferred spawn、attach level_container、prop_data→MonsterBase.data。",
        "【層級】ForegroundCanopyHoist 子改掛 LevelContainer；Phase10 表列與層級規則已去 WallFill 為主敘述。",
        "【戰鬥】perform_ghost_dash 扇形掃向＋move_and_collide 防撞。",
        "【換關】UILayer 不隨 LevelContainer 換掉；黑幕→換子場景→request_area_title／set_player_in_homestead；採收鈕顯隱須狀態機對齊（待修列 §5）。",
        "【下一包】家園 HomesteadLevel 全圖、動畫粒子、寵物碰撞／HarvestToggle／搖桿／NPC道具影子。",
    ]
    for i, bullet in enumerate(bullets, start=1):
        ws02.cell(r02 + i, 3, bullet)

    ws03 = wb[s03]
    note_1001 = ws03.cell(1001, 6).value or ""
    extra_map = " LakeSideLevel 地圖定版 2026-03-29（ARCHITECTURE 湖畔章＋下一階段第0項）。"
    if isinstance(note_1001, str) and "地圖定版 2026-03-29" not in note_1001:
        ws03.cell(1001, 6, (note_1001 + extra_map).strip())

    r03 = last_used_row(ws03) + 1
    ws03.cell(r03, 1, "P0")
    ws03.cell(
        r03,
        2,
        "湖畔 LakeSideLevel 大地圖＋碰撞／撒点／前景 hoist 定版（2026-03-29）",
    )
    ws03.cell(
        r03,
        3,
        "ARCHITECTURE 新章完整；下一工作包：動畫粒子、家園獨立場景全圖、§5 四項 bug、作物嘟嘟時間軸",
    )
    ws03.cell(
        r03,
        4,
        "LakeSideLevel.tscn／LakeSideLevelRoot／MarkersPropSpawner／ForegroundCanopyHoist／MonsterBase.perform_ghost_dash",
    )
    ws03.cell(
        r03,
        5,
        "換關不捨 UILayer 儀式 UI；Main LevelContainer 換子場景；HomeManager.request_area_title",
    )
    ws03.cell(
        r03,
        6,
        "04 數值無變更；DevLog 本列為追加不覆寫舊 P0/P1 列。",
    )

    ws07 = wb[s07]
    r07 = last_used_row(ws07) + 1
    ws07.cell(r07, 1, SYNC_DATE)
    ws07.cell(r07, 2, "Phase 9＋湖畔地圖／文件")
    ws07.cell(
        r07,
        3,
        "ARCHITECTURE 湖畔關卡地圖定版章：碰撞分段、撒点 deferred、前景 y_sort、奧義防撞、換關儀式感協議、家園下一張注意事項；"
        "下一階段第0項改為 polish＋家園全圖；常見地雷增 busy parent add_child。",
    )
    ws07.cell(
        r07,
        5,
        "待修隊列：寵物地形、對話後採收鈕、進出家園搖桿、NPC／道具影子、場景動畫粒子。"
        "04 數值無變更。快照見 docs/_ARCHITECTURE_sync_snapshot_for_devlog.txt。",
    )

    ws100 = wb[s100]
    old_title = ws100.cell(1002, 2).value
    if old_title and "架構預研" in str(old_title):
        ws100.cell(
            1002,
            2,
            "Phase9：NPC 對話互動（MVP 已落地；表驅動／多 NPC 待延伸）",
        )
    proposal_1002 = ws100.cell(1002, 5).value or ""
    mvp_line = (
        "MVP：LakeSideLevel＋lakeside_smith、DialogueManager／NpcInteractionManager、"
        "inventory_grant_requested→InventoryManager、DialogueHudLocker／z_index／搖桿 process_input 對齊。"
    )
    if isinstance(proposal_1002, str) and "MVP：" not in proposal_1002:
        ws100.cell(1002, 5, (proposal_1002 + " " + mvp_line).strip())
    ws100.cell(1002, 8, "已落地（MVP）")

    r100 = last_used_row(ws100) + 1
    ws100.cell(r100, 1, SYNC_DATE.isoformat())
    ws100.cell(r100, 2, "湖畔地圖段落結案後：動畫粒子＋儀式 UI bug 對齊＋家園全圖")
    ws100.cell(r100, 3, "ARCHITECTURE 湖畔章 §5＋下一階段第0項")
    ws100.cell(
        r100,
        4,
        "水／花／螢火蟲／火把動畫與粒子；寵物 CharacterBody 撞世界層；HarvestToggle 與 in_homestead／對話關閉邊界；搖桿 show 與 process_input；ShadowComponent 道具與 NPC",
    )
    ws100.cell(
        r100,
        5,
        "場景內 AnimatedSprite2D／CPUParticles2D 或既有 EffectManager；HomeManager 狀態堆疊復盤；家園 HomesteadLevel 換 LevelContainer 子實例",
    )
    ws100.cell(r100, 6, "與 DialogueHudLocker／HarvestHudLocker 互斥需單一真相避免回歸")
    ws100.cell(r100, 7, "村外對話後採收鈕、進出家園搖桿、寵物穿牆三項優先復現")
    ws100.cell(r100, 8, "待驗證")

    r100b = last_used_row(ws100) + 1
    ws100.cell(r100b, 1, SYNC_DATE.isoformat())
    ws100.cell(r100b, 2, "Phase 9：對話框帳簿風與左右欄幾何對齊（仍排隊）")
    ws100.cell(r100b, 3, "ARCHITECTURE「下一小輪」")
    ws100.cell(
        r100b,
        4,
        "左欄仍灰底咖啡字無框；與右側選項長條高度／上下緣視覺不一致",
    )
    ws100.cell(
        r100b,
        5,
        "左 PanelContainer 套用與 PetUI 一致 ledger StyleBoxFlat；必要時外層 Margin／HBox min 高度同步",
    )
    ws100.cell(r100b, 6, "小螢幕換行溢出")
    ws100.cell(r100b, 7, "360×640 實機預覽一輪對話三選項")
    ws100.cell(r100b, 8, "待驗證")

    wb.save(XLSX)
    print("Saved:", XLSX)
    print("Snapshot:", docs_snap)
    print("02 new block start row:", r02)
    print("03 new row:", r03)
    print("07 new row:", r07)
    print("100 new rows:", r100, r100b)


if __name__ == "__main__":
    main()
