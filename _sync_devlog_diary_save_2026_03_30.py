# -*- coding: utf-8 -*-
"""Append ARCHITECTURE Phase 11（日記／單槽存檔）to Monster_DevLog_v4.xlsx (02/03/07/100) + docs snapshot.

若專案根目錄沒有 Monster_DevLog_v4.xlsx，仍會寫入 docs/_ARCHITECTURE_sync_snapshot_for_devlog.txt，
並印出提示請將試算表放回根目錄後再執行。
"""
from __future__ import annotations

import datetime
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Monster_DevLog_v4.xlsx"
ARCH = ROOT / "ARCHITECTURE.md"
SYNC_DATE = datetime.date(2026, 3, 30)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def arch_toc(md: str) -> str:
    lines = md.splitlines()
    toc_lines = []
    for i, ln in enumerate(lines, 1):
        if ln.startswith("## "):
            toc_lines.append(f"L{i}: {ln[3:].strip()}")
    return "\n".join(toc_lines)


def extract_phase11_block(md: str) -> str:
    lines = md.splitlines()
    out: list[str] = []
    capture = False
    for ln in lines:
        if ln.startswith("## Phase 11："):
            capture = True
        elif capture and ln.startswith("## ") and "Phase 11" not in ln:
            break
        if capture:
            out.append(ln)
    return "\n".join(out) if out else "(未找到 Phase 11 區塊)"


def find_sheet(wb, prefix: str) -> str:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    raise KeyError(prefix)


def last_used_row(ws, max_col: int = 12) -> int:
    last = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None and v != "":
                last = r
                break
    return last


def main() -> None:
    import openpyxl

    md = ARCH.read_text(encoding="utf-8")
    toc = arch_toc(md)
    p11 = extract_phase11_block(md)
    snapshot = (
        f"《怪物與我》ARCHITECTURE.md 同步快照 {SYNC_DATE.isoformat()}（Phase 11 日記／單槽存檔）\n\n"
        f"【章節索引】\n{toc}\n\n---\n{p11}\n\n"
        f"（全文見 repo 根目錄 ARCHITECTURE.md；SignalBus game_save_requested／finished；"
        f"存檔路徑 user://monster_and_i_save_v1.json；04 數值中心本次無變更。）"
    )
    if len(snapshot) > 32000:
        snapshot = snapshot[:31900] + "\n…(截斷)"

    docs_snap = ROOT / "docs" / "_ARCHITECTURE_sync_snapshot_for_devlog.txt"
    docs_snap.parent.mkdir(parents=True, exist_ok=True)
    docs_snap.write_text(snapshot, encoding="utf-8")
    print("Snapshot:", docs_snap)

    if not XLSX.is_file():
        print(
            "WARN: 找不到", XLSX.name,
            "— 已僅更新 docs 快照。請將試算表置於專案根目錄後再執行本腳本以寫入 02/03/07/100。",
        )
        return

    wb = openpyxl.load_workbook(XLSX)
    s02 = find_sheet(wb, "02_")
    s03 = find_sheet(wb, "03_")
    s07 = find_sheet(wb, "07_")
    s100 = find_sheet(wb, "100_")

    summary = (
        "Phase 11 落地：DiaryManager（心情筆記遞增＋生涯成就 CAREER_TITLES）；"
        "SaveGameManager 單槽 JSON v1（NpcState／Inventory／Pet／Diary／關卡／玩家 HP 座標）；"
        "Main.gd await 讀檔；DialogueEffectEntry.career_milestone_id＋lakeside_smith_graph 範例；"
        "DiaryUI 分頁對齊 Inventory tab；SaveGameButton／SaveProgressOverlay／game_save_*；"
        "HomeManager switch_*_async、set_player_in_homestead(show_banner)；SealHudLocker 含存檔鈕。"
    )

    r02 = last_used_row(wb[s02]) + 1
    ws02 = wb[s02]
    ws02.cell(r02, 1, SYNC_DATE.isoformat())
    ws02.cell(r02, 2, "ARCHITECTURE Phase 11 日記／單槽存檔")
    ws02.cell(r02, 3, summary)
    ws02.cell(
        r02,
        4,
        "DiaryManager；SaveGameManager；game_save_*；career_milestone_id；Main.gd；SealHudLocker SaveGameButton",
    )
    ws02.cell(r02, 5, "是")
    bullets = [
        "【日記】心情筆記 append＋捲到底；生涯 try_unlock_career；對話 GIVE_ITEM 後解鎖 milestone。",
        "【存檔】user://monster_and_i_save_v1.json；有存檔則略過 Inventory／Pet 開局種子。",
        "【訊號】game_save_requested／finished；Overlay 最短顯示時間；寫檔與讀檔 await 換關。",
        "【釐清】寵物與主角同路徑進 Die→_spawn_loot；掉落感與 drop_chance 機率有關。",
        "【檔案】docs/_ARCHITECTURE_sync_snapshot_for_devlog.txt 已更新。",
    ]
    for i, b in enumerate(bullets, start=1):
        ws02.cell(r02 + i, 3, b)

    r03 = last_used_row(wb[s03]) + 1
    ws03 = wb[s03]
    ws03.cell(r03, 1, "P0")
    ws03.cell(r03, 2, "日記系統＋單槽存檔（Phase 11，2026-03-30）")
    ws03.cell(
        r03,
        3,
        "心情／生涯分頁、對話掛成就 id、本機覆寫存檔、啟動還原、NpcState 納管",
    )
    ws03.cell(
        r03,
        4,
        "ARCHITECTURE.md；DiaryUI；SaveGameManager；project.godot autoload；lakeside_smith_graph.tres",
    )
    ws03.cell(
        r03,
        5,
        "後續：更多 career_milestone_id、存檔 version 遷移、雲端（未定）",
    )
    ws03.cell(r03, 6, "04 無本批數值變更")

    r07 = last_used_row(wb[s07]) + 1
    ws07 = wb[s07]
    ws07.cell(r07, 1, SYNC_DATE)
    ws07.cell(r07, 2, "Phase 11 日記／存檔")
    ws07.cell(
        r07,
        3,
        "ARCHITECTURE 新增 Phase 11、SignalBus 表、目錄 autoload、下一階段與靈感牆更新；"
        "DevLog 02／03／07／100 本批附加；快照更新。",
    )
    ws07.cell(r07, 5, "04 無數值變更。")

    r100 = last_used_row(wb[s100]) + 1
    ws100 = wb[s100]
    ws100.cell(r100, 1, SYNC_DATE.isoformat())
    ws100.cell(r100, 2, "生涯成就擴充：僅增 DialogueEffectEntry 欄位＋ DiaryManager 標題表")
    ws100.cell(r100, 3, f"ARCHITECTURE 同步 {SYNC_DATE.isoformat()}")
    ws100.cell(
        r100,
        4,
        "新怪物／任務解鎖＝新 milestone id＋ CAREER_TITLES 登記；避免 SignalBus 塞邏輯",
    )
    ws100.cell(
        r100,
        5,
        "與 inventory_grant、grant_once 同一套對話效果管線",
    )
    ws100.cell(r100, 6, "—")
    ws100.cell(r100, 7, "—")
    ws100.cell(r100, 8, "已落地")

    wb.save(XLSX)
    print("Saved:", XLSX)
    print("02 block start:", r02)
    print("03 row:", r03)
    print("07 row:", r07)
    print("100 row:", r100)


if __name__ == "__main__":
    main()
