# -*- coding: utf-8 -*-
"""Append ARCHITECTURE／Phase10 里程碑 to Monster_DevLog_v4.xlsx (02/03/07/100) + docs snapshot."""
from __future__ import annotations

import datetime
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Monster_DevLog_v4.xlsx"
ARCH = ROOT / "ARCHITECTURE.md"

SYNC_DATE = datetime.date(2026, 3, 29)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def arch_toc(md: str) -> str:
    lines = md.splitlines()
    toc_lines = []
    for i, ln in enumerate(lines, 1):
        if ln.startswith("## "):
            toc_lines.append(f"L{i}: {ln[3:].strip()}")
    return "\n".join(toc_lines)


def extract_phase10_block(md: str) -> str:
    lines = md.splitlines()
    out: list[str] = []
    capture = False
    for ln in lines:
        if ln.startswith("## Phase 10：家園與採收"):
            capture = True
        elif capture and ln.startswith("## ") and "Phase 10" not in ln:
            break
        if capture:
            out.append(ln)
    return "\n".join(out) if out else "(未找到 Phase 10 區塊)"


def find_sheet(wb: openpyxl.Workbook, prefix: str) -> str:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    raise KeyError(prefix)


def last_used_row(ws: openpyxl.worksheet.worksheet.Worksheet, max_col: int = 12) -> int:
    last = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None and v != "":
                last = r
                break
    return last


def main() -> None:
    md = ARCH.read_text(encoding="utf-8")
    toc = arch_toc(md)
    p10 = extract_phase10_block(md)
    snapshot = (
        f"《怪物與我》ARCHITECTURE.md 同步快照 {SYNC_DATE.isoformat()}（Phase10＋世界提示／payload／打字漸隱）\n\n"
        f"【章節索引】\n{toc}\n\n---\n{p10}\n\n"
        f"（全文見 repo 根目錄 ARCHITECTURE.md；player_world_hint_changed 第三參數必傳 null 或 Dictionary；"
        f"「下一階段」第 5 項已補世界提示擴充留線。）"
    )
    if len(snapshot) > 32000:
        snapshot = snapshot[:31900] + "\n…(截斷)"

    docs_snap = ROOT / "docs" / "_ARCHITECTURE_sync_snapshot_for_devlog.txt"
    docs_snap.parent.mkdir(parents=True, exist_ok=True)
    docs_snap.write_text(snapshot, encoding="utf-8")

    wb = openpyxl.load_workbook(XLSX)
    s02 = find_sheet(wb, "02_")
    s03 = find_sheet(wb, "03_")
    s07 = find_sheet(wb, "07_")
    s100 = find_sheet(wb, "100_")

    summary = (
        "世界提示系統落地：SignalBus player_world_hint_changed(hint_id,show_hint,payload)（signal 無預設值，"
        "無 payload 則 emit null）；PlayerHintCatalog；HarvestModeHint 打字序列、final_hold／fade、_seq_token 打斷；"
        "HomeManager 依 counts_as_mature_available 切換家園採收教學、採光後 payload 收工＋自動關採收；"
        "HomesteadCrop counts_as_mature_available；ARCHITECTURE Phase10 表／訊號／留線／下一階段第5項補述。"
    )

    r02 = last_used_row(wb[s02]) + 1
    ws02 = wb[s02]
    ws02.cell(r02, 1, SYNC_DATE.isoformat())
    ws02.cell(r02, 2, "ARCHITECTURE 世界提示／家園採收教學")
    ws02.cell(r02, 3, summary)
    ws02.cell(
        r02,
        4,
        "player_world_hint_changed；HarvestModeHint；HomeManager _sync；HomesteadCrop；留線寶箱／危險擴充",
    )
    ws02.cell(r02, 5, "是")
    bullets = [
        "【訊號】player_world_hint_changed 三參數；payload Dictionary：typing_intro／final_text／timing／final_fade_out_sec 等。",
        "【邏輯】有成熟可採才「點採收」；採收中「拖曳」；採光後打字收工＋關採收；item_collected 刷新計數。",
        "【技術】已採株用 counts_as_mature_available 排除 _gathered 誤計；離園 emit(\"\", false, null)。",
        "【擴充】寶箱／危險／教學＝新 hint_id 或 payload，原則不重複宣告訊號（見聖經留線）。",
        "【檔案】docs/_ARCHITECTURE_sync_snapshot_for_devlog.txt 已更新。",
    ]
    for i, b in enumerate(bullets, start=1):
        ws02.cell(r02 + i, 3, b)

    r03 = last_used_row(wb[s03]) + 1
    ws03 = wb[s03]
    ws03.cell(r03, 1, "P0")
    ws03.cell(r03, 2, "世界提示（主角頭上情境提示）＋家園採收教學 UX")
    ws03.cell(
        r03,
        3,
        "payload 打字／漸隱、條件提示、自動關採收已實作；聖經與 DevLog 同步",
    )
    ws03.cell(r03, 4, "ARCHITECTURE.md；SignalBus；HomeManager；HarvestModeHint；PlayerHintCatalog；HomesteadCrop")
    ws03.cell(
        r03,
        5,
        "後續：寶箱／危險觸發端；翻譯表／.csv 若與 PlayerHintCatalog 分流",
    )
    ws03.cell(r03, 6, "04 無本批數值變更")

    r07 = last_used_row(wb[s07]) + 1
    ws07 = wb[s07]
    ws07.cell(r07, 1, SYNC_DATE)
    ws07.cell(r07, 2, "架構／世界提示")
    ws07.cell(
        r07,
        3,
        "ARCHITECTURE 更新 Phase10 落地表、訊號細節、留線、下一階段第5項；DevLog 02／03／07／100 附加；快照更新。",
    )
    ws07.cell(r07, 5, "04 無數值變更。")

    r100 = last_used_row(wb[s100]) + 1
    ws100 = wb[s100]
    ws100.cell(r100, 1, SYNC_DATE.isoformat())
    ws100.cell(r100, 2, "世界提示：可擴充單訊號多型態演出")
    ws100.cell(r100, 3, f"ARCHITECTURE 同步 {SYNC_DATE.isoformat()}")
    ws100.cell(
        r100,
        4,
        "採收教學為首個完整案例；payload 供打字＋淡出；後續系統共用 HarvestModeHint",
    )
    ws100.cell(
        r100,
        5,
        "禁止 SignalBus 寫邏輯；新情境優先 hint_id／payload 與 HarvestModeHint 約定",
    )
    ws100.cell(r100, 6, "—")
    ws100.cell(r100, 7, "—")
    ws100.cell(r100, 8, "已入聖經；可平行擴充")

    wb.save(XLSX)
    print("Saved:", XLSX)
    print("Snapshot:", docs_snap)
    print("02 block start:", r02)
    print("03 row:", r03)
    print("07 row:", r07)
    print("100 row:", r100)


if __name__ == "__main__":
    main()
