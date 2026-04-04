# -*- coding: utf-8 -*-
"""ARCHITECTURE.md 待辦收口（儀式感 UI／Phase8／頭飾／底欄／背包頭飾 UI 等標已結案）→ DevLog 02/03/07/100 + docs 快照一句話。"""
from __future__ import annotations

import datetime
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Monster_DevLog_v4.xlsx"
ARCH = ROOT / "ARCHITECTURE.md"
SYNC_DATE = datetime.date(2026, 3, 30)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def arch_toc_one_line(md: str) -> str:
    n = 0
    for ln in md.splitlines():
        if ln.startswith("## "):
            n += 1
    return f"ARCHITECTURE.md 共偵測 {n} 個 ## 章節（詳見檔案）"


def find_sheet(wb: openpyxl.Workbook, prefix: str) -> str:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    raise KeyError(prefix)


def last_used_row(ws, max_col: int = 12) -> int:
    last = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None and v != "":
                last = r
                break
    return last


def main() -> None:
    md = ARCH.read_text(encoding="utf-8")
    one = arch_toc_one_line(md)
    snap = (
        f"《怪物與我》{SYNC_DATE.isoformat()}：ARCHITECTURE 待辦收口— "
        "湖畔§5 儀式感（寵物碰撞／採收鈕／搖桿）、Phase9 帳簿風、Phase8／底欄／頭飾運營／背包道具與頭飾請求 UI "
        "改标已结或维护备注；下一階段清單重編 0～5。"
        f" {one}"
    )
    docs = ROOT / "docs" / "_ARCHITECTURE_sync_snapshot_for_devlog.txt"
    prev = docs.read_text(encoding="utf-8") if docs.is_file() else ""
    docs.parent.mkdir(parents=True, exist_ok=True)
    docs.write_text(prev + "\n\n---\n\n" + snap, encoding="utf-8")
    print("Appended:", docs)

    if not XLSX.is_file():
        print("WARN: 找不到", XLSX.name)
        return

    wb = openpyxl.load_workbook(XLSX)
    s02 = find_sheet(wb, "02_")
    s03 = find_sheet(wb, "03_")
    s07 = find_sheet(wb, "07_")
    s100 = find_sheet(wb, "100_")

    summary = (
        "ARCHITECTURE 收口：儀式感 UI 三項改「已結案」；"
        "Phase8／三面板 UI_BOTTOM_BAR_HEIGHT_PX／頭飾 offset 慣例／背包道具與頭飾請求合併為已結案（2026-03）；"
        "下一階段 0～5 重編；§5 採集物陰影與場景粒子列為非阻塞 polish。"
    )

    r02 = last_used_row(wb[s02]) + 1
    ws02 = wb[s02]
    ws02.cell(r02, 1, SYNC_DATE.isoformat())
    ws02.cell(r02, 2, "ARCHITECTURE 待辦收口（與程式現況對齊）")
    ws02.cell(r02, 3, summary)
    ws02.cell(r02, 4, "ARCHITECTURE.md 湖畔§5；下一階段；Phase 9 狀態列")
    ws02.cell(r02, 5, "是")
    for i, b in enumerate(
        [
            "【文件】舊「已知待修」中寵物／採收鈕／搖桿—已結案敘述。",
            "【文件】Phase 9 帳簿風、下一小輪區塊改已併入實作。",
            "【文件】Phase 8／頭飾運營／底欄／背包頭飾—合併為單條已結案＋新美資產驗收指引。",
            "【總誌】本批與 03 舊 P0「HarvestToggle／對話」並存時以 ARCHITECTURE 為準。",
        ],
        start=1,
    ):
        ws02.cell(r02 + i, 3, b)

    r03 = last_used_row(wb[s03]) + 1
    ws03 = wb[s03]
    ws03.cell(r03, 1, "P2")
    ws03.cell(r03, 2, "總誌佇列與 ARCHITECTURE「下一階段」對齊（2026-03-30）")
    ws03.cell(
        r03,
        3,
        "儀式感 UI、三面板、頭飾與背包請求—文件已收口；"
        "03 表內若有舊「待修採收鈕／搖桿」列可標「已結案」或不再拉升優先。",
    )
    ws03.cell(r03, 4, "Monster_DevLog_v4.xlsx 03_；ARCHITECTURE.md")
    ws03.cell(r03, 5, "新願景（環境生物／進化／指引／NPC×寵物）見對話紀錄，尚未升格任務")
    ws03.cell(r03, 6, "04 本次無強制數值表更新")

    r07 = last_used_row(wb[s07]) + 1
    ws07 = wb[s07]
    ws07.cell(r07, 1, SYNC_DATE)
    ws07.cell(r07, 2, "文件維護：ARCHITECTURE 待辦淨空舊項")
    ws07.cell(
        r07,
        3,
        "與團隊實作與玩家驗收口徑一致；減少新對話誤以為採收鈕／搖桿仍為未修 bug。",
    )
    ws07.cell(r07, 5, snap[:500])

    r100 = last_used_row(wb[s100]) + 1
    ws100 = wb[s100]
    ws100.cell(r100, 1, SYNC_DATE.isoformat())
    ws100.cell(r100, 2, "願景預研（僅記錄）：非戰鬥可封印生物／寵物進化／邊緣指引箭／NPC×寵物演出")
    ws100.cell(r100, 3, "ARCHITECTURE 對齊後討論；見 2026-03-30 對話—未升格 Phase")
    ws100.cell(r100, 4, "Ambient 與 monsters 群組分岔、SealManager 目標篩選、buff 堆疊與存檔")
    ws100.cell(r100, 5, "Evolution：PetResource 切換＋XP 表；ObjectiveHint UI 與 minimap 邊緣")
    ws100.cell(r100, 6, "—")
    ws100.cell(r100, 7, "—")
    ws100.cell(r100, 8, "未定案")

    wb.save(XLSX)
    print("Saved:", XLSX)
    print("02:", r02, "03:", r03, "07:", r07, "100:", r100)


if __name__ == "__main__":
    main()
